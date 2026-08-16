"""
Haftalık iş takip Excel'ini güvenli biçimde parse etme ve rapor üretme.

Önemli ayrım:
- Bu modül önce Excel'in tamamını parse + validate eder.
- Hiçbir global state'i parse sırasında değiştirmez.
- Tüm satırlar geçerliyse depo.excel_import_uygula(...) çağrılır.
"""

from __future__ import annotations

import io
import os
import re
from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import depo


class ExcelAktarimHatasi(Exception):
    pass


TURKCE_HARFLER = str.maketrans(
    {
        "ç": "c",
        "Ç": "C",
        "ğ": "g",
        "Ğ": "G",
        "ı": "i",
        "İ": "I",
        "ö": "o",
        "Ö": "O",
        "ş": "s",
        "Ş": "S",
        "ü": "u",
        "Ü": "U",
        "â": "a",
        "Â": "A",
    }
)


def _sadelestir(deger):
    """'Kart Üretim Adet' -> 'KART URETIM ADET'."""
    metin = str(deger or "").translate(TURKCE_HARFLER).upper()
    metin = metin.replace(".", " ").replace("_", " ")
    return re.sub(r"\s+", " ", metin).strip()


BASLIK_ESLESME = {
    _sadelestir(k): v
    for k, v in {
        "Sıra": "sira",
        "Talep NO": "talep_no",
        "Talep Sahibi": "talep_sahibi",
        "Kart Stok No": "stok_no",
        "Kart Üretim Adet": "adet_metin",
        "Üretim Adet": "adet_metin",
        "Adet": "adet_metin",
        "Planlanan Başlangıç T.": "plan_hafta",
        "Dizgi Başlama Tarihi": "plan_baslama",
        "Planlanan Teslim T.": "plan_teslim",
        "Gerçekleşen Teslim T.": "gerceklesen_teslim",
        "DURUM": "excel_durum",
        "PCB": "pcb",
    }.items()
}

# Substring kontrolü yok. Yalnız exact normalize edilmiş değerler eşleşir.
TESLIM_DURUMLARI = {
    "TESLIM",
    "TESLIM EDILDI",
    "TESLIM EDILMISTIR",
    "TAMAMLANDI",
}
DEVAM_DURUMLARI = {
    "DIZGI",
    "DIZGIDE",
    "DIZGIYE ALINDI",
    "DIZGIYE VERILDI",
    "DEVAM EDIYOR",
}
SIRADA_DURUMLARI = {
    "",
    "SIRADA",
    "BEKLIYOR",
    "BEKLEMEDE",
    "TESLIM EDILMEDI",
}


def durum_coz(deger):
    """
    Excel kaynak durumunu internal başlangıç durumuna çevirir.

    Dönüş: (durum, uyari_var_mi)
    Bilinmeyen durum kartı yanlışlıkla TAMAM yapmaz; SIRADA başlatır ve uyarı sayar.
    """
    temiz = _sadelestir(deger)
    if temiz in TESLIM_DURUMLARI:
        return depo.TAMAM, False
    if temiz in DEVAM_DURUMLARI:
        return depo.DEVAM, False
    if temiz in SIRADA_DURUMLARI:
        return depo.SIRADA, False
    return depo.SIRADA, True


def adet_coz(deger):
    """Pozitif üretim adedini çözer; sessizce 1'e düşmek yerine hatalı değeri reddeder."""
    if deger is None or str(deger).strip() == "":
        raise ExcelAktarimHatasi("Üretim adedi boş olamaz.")

    if isinstance(deger, bool):
        raise ExcelAktarimHatasi("Üretim adedi sayı olmalı.")

    if isinstance(deger, (int, float)):
        if isinstance(deger, float) and not deger.is_integer():
            raise ExcelAktarimHatasi(f"Üretim adedi tam sayı olmalı: {deger}")
        adet = int(deger)
        if adet < 1:
            raise ExcelAktarimHatasi(f"Üretim adedi en az 1 olmalı: {deger}")
        return adet

    metin = str(deger).strip()
    esle = re.search(r"\d[\d.,\s]*", metin)
    if not esle:
        raise ExcelAktarimHatasi(f"Üretim adedi okunamadı: '{metin}'")

    sayi = re.sub(r"[\s.,]", "", esle.group())
    if not sayi.isdigit():
        raise ExcelAktarimHatasi(f"Üretim adedi okunamadı: '{metin}'")
    adet = int(sayi)
    if adet < 1:
        raise ExcelAktarimHatasi(f"Üretim adedi en az 1 olmalı: '{metin}'")
    return adet


def _baslik_satiri_bul(ws):
    for i, satir in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
        temiz = [_sadelestir(h) for h in satir]
        if "TALEP NO" in temiz or "KART STOK NO" in temiz:
            return i, temiz
    return None, None


def _hucre_al(satir, kolonlar, alan):
    i = kolonlar.get(alan)
    if i is None or i >= len(satir):
        return None
    deger = satir[i]
    return deger if deger != "" else None


def excelden_aktar(dosya_yolu, kullanici):
    """
    Haftalık Excel'i önce tamamen parse eder ve doğrular; sonra tek commit çağrısı yapar.

    Source-of-truth politikası:
    - Excel: planlama + kaynak durum bilgisi
    - Web uygulaması: mevcut kartların operasyonel workflow durumu
    - Yeni kart: Excel kaynak durumu yalnız ilk internal durumu belirlemek için kullanılır
    """
    try:
        wb = openpyxl.load_workbook(dosya_yolu, data_only=True, read_only=True)
    except Exception as exc:  # noqa: BLE001
        raise ExcelAktarimHatasi(f"Dosya geçerli bir Excel çalışma kitabı değil: {exc}") from exc

    try:
        ws = wb[wb.sheetnames[0]]
        baslik_no, basliklar = _baslik_satiri_bul(ws)
        if not baslik_no:
            raise ExcelAktarimHatasi(
                "Başlık satırı bulunamadı. Dosyada 'Talep NO' veya 'Kart Stok No' sütunu olmalı."
            )

        kolonlar = {}
        for idx, baslik in enumerate(basliklar):
            alan = BASLIK_ESLESME.get(baslik)
            if alan and alan not in kolonlar:
                kolonlar[alan] = idx

        if "stok_no" not in kolonlar and "talep_no" not in kolonlar:
            raise ExcelAktarimHatasi("'Talep NO' ya da 'Kart Stok No' sütunu okunamadı.")
        if "adet_metin" not in kolonlar:
            raise ExcelAktarimHatasi(
                "Üretim adedi sütunu bulunamadı. 'Kart Üretim Adet', 'Üretim Adet' veya 'Adet' olmalı."
            )

        parsed = []
        gorulen_anahtarlar = set()
        uyari_sayisi = 0

        for excel_satir_no, satir in enumerate(
            ws.iter_rows(min_row=baslik_no + 1, values_only=True),
            start=baslik_no + 1,
        ):
            if not any(h not in (None, "") for h in satir):
                continue

            talep_no = str(_hucre_al(satir, kolonlar, "talep_no") or "").strip()
            stok_no = str(_hucre_al(satir, kolonlar, "stok_no") or "").strip()
            if not talep_no and not stok_no:
                continue

            # Mevcut business key korunuyor; duplicate varsa sessiz merge yerine import reddedilir.
            anahtar = f"{talep_no}|{stok_no}"
            if anahtar in gorulen_anahtarlar:
                raise ExcelAktarimHatasi(
                    f"Excel satır {excel_satir_no}: aynı Talep NO + Kart Stok No kombinasyonu "
                    f"birden fazla kez bulundu ({anahtar}). Kaynakta benzersiz anahtar gerekli."
                )
            gorulen_anahtarlar.add(anahtar)

            try:
                toplam_adet = adet_coz(_hucre_al(satir, kolonlar, "adet_metin"))
            except ExcelAktarimHatasi as exc:
                raise ExcelAktarimHatasi(f"Excel satır {excel_satir_no}: {exc}") from exc

            plan_baslama_ham = _hucre_al(satir, kolonlar, "plan_baslama")
            plan_teslim_ham = _hucre_al(satir, kolonlar, "plan_teslim")
            gerceklesen_ham = _hucre_al(satir, kolonlar, "gerceklesen_teslim")

            plan_baslama = depo.tarih_coz(plan_baslama_ham)
            plan_teslim = depo.tarih_coz(plan_teslim_ham)
            gerceklesen = depo.tarih_coz(gerceklesen_ham)

            if plan_baslama_ham not in (None, "") and not plan_baslama:
                raise ExcelAktarimHatasi(
                    f"Excel satır {excel_satir_no}: Dizgi Başlama Tarihi okunamadı."
                )
            if plan_teslim_ham not in (None, "") and not plan_teslim:
                raise ExcelAktarimHatasi(
                    f"Excel satır {excel_satir_no}: Planlanan Teslim Tarihi okunamadı."
                )
            if gerceklesen_ham not in (None, "") and not gerceklesen:
                raise ExcelAktarimHatasi(
                    f"Excel satır {excel_satir_no}: Gerçekleşen Teslim Tarihi okunamadı."
                )
            if plan_baslama and plan_teslim and plan_baslama > plan_teslim:
                raise ExcelAktarimHatasi(
                    f"Excel satır {excel_satir_no}: Plan başlangıç tarihi plan teslim tarihinden sonra olamaz."
                )

            excel_durum_raw = _hucre_al(satir, kolonlar, "excel_durum")
            ilk_durum, uyari = durum_coz(excel_durum_raw)
            if uyari:
                uyari_sayisi += 1

            sira_raw = _hucre_al(satir, kolonlar, "sira")
            if isinstance(sira_raw, (int, float)):
                sira = int(sira_raw)
            elif sira_raw in (None, ""):
                sira = None
            else:
                try:
                    sira = int(str(sira_raw).strip())
                except ValueError:
                    sira = None
                    uyari_sayisi += 1

            plan = {
                "sira": sira,
                "talep_sahibi": str(_hucre_al(satir, kolonlar, "talep_sahibi") or "").strip(),
                "toplam_adet": toplam_adet,
                "adet_metin": str(_hucre_al(satir, kolonlar, "adet_metin") or "").strip(),
                "plan_hafta": str(_hucre_al(satir, kolonlar, "plan_hafta") or "").strip(),
                "plan_baslama": plan_baslama,
                "plan_teslim": plan_teslim,
                "excel_durum": str(excel_durum_raw or "").strip(),
                "pcb": str(_hucre_al(satir, kolonlar, "pcb") or "").strip(),
            }

            parsed.append(
                {
                    "anahtar": anahtar,
                    "talep_no": talep_no,
                    "stok_no": stok_no,
                    "plan": plan,
                    "gerceklesen_teslim": gerceklesen,
                    "ilk_durum": ilk_durum,
                }
            )

        if not parsed:
            raise ExcelAktarimHatasi("Excel'de işlenecek kart satırı bulunamadı.")

    finally:
        wb.close()

    return depo.excel_import_uygula(
        dosya_adi=os.path.basename(dosya_yolu),
        kullanici=kullanici,
        satirlar=parsed,
        uyari_sayisi=uyari_sayisi,
    )


# ---------------------------------------------------------------- rapor üretimi

BASLIK_DOLGU = PatternFill("solid", fgColor="0F2027")
BASLIK_YAZI = Font(name="Arial", bold=True, color="FFFFFF", size=11)
GOVDE_YAZI = Font(name="Arial", size=10)


def _sayfa_yaz(ws, basliklar, satirlar):
    ws.append(basliklar)
    for hucre in ws[1]:
        hucre.fill = BASLIK_DOLGU
        hucre.font = BASLIK_YAZI
        hucre.alignment = Alignment(horizontal="center", vertical="center")
    for satir in satirlar:
        ws.append(satir)
    for sutun in range(1, len(basliklar) + 1):
        en = max(
            [len(str(basliklar[sutun - 1]))]
            + [len(str(s[sutun - 1])) for s in satirlar[:400]]
            or [10]
        )
        ws.column_dimensions[get_column_letter(sutun)].width = min(38, max(12, en + 3))
    ws.freeze_panes = "A2"
    for satir in ws.iter_rows(min_row=2):
        for hucre in satir:
            hucre.font = GOVDE_YAZI


def calisma_kitabi_uret(kartlar, loglar, ozet):
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "Kart Durumları"
    _sayfa_yaz(
        ws,
        [
            "Sıra",
            "Talep NO",
            "Talep Sahibi",
            "Kart Stok No",
            "Toplam Adet",
            "Tamamlanan",
            "Kalan",
            "Durum",
            "Kaynak Durumu",
            "Değerlendirme",
            "Sapma (gün)",
            "Plan Başlangıç",
            "Plan Teslim",
            "Başlama Zamanı",
            "Bitiş Zamanı",
            "Operatör",
            "Not",
            "PCB",
            "Kaynakta Aktif",
            "Admin Gizli",
        ],
        [
            [
                k.get("sira"),
                k.get("talep_no"),
                k.get("talep_sahibi"),
                k.get("stok_no"),
                k.get("toplam_adet"),
                k.get("tamamlanan_adet"),
                k.get("kalan_adet"),
                k.get("durum"),
                k.get("excel_durum") or "",
                k.get("rozet"),
                k.get("sapma") if k.get("sapma") is not None else "",
                k.get("plan_baslama") or "",
                k.get("plan_teslim") or "",
                k.get("baslama_zamani") or "",
                k.get("bitis_zamani") or "",
                k.get("operator") or "",
                k.get("aciklama") or "",
                k.get("pcb") or "",
                k.get("source_active", 1),
                k.get("admin_gizli", 0),
            ]
            for k in kartlar
        ],
    )

    ws2 = wb.create_sheet("İşlem Logu")
    _sayfa_yaz(
        ws2,
        ["Zaman", "Kullanıcı", "Rol", "İşlem", "Talep NO", "Kart Stok No", "Adet", "Detay"],
        [
            [
                l.get("zaman"),
                l.get("kullanici"),
                l.get("rol"),
                l.get("islem"),
                l.get("talep_no") or "",
                l.get("stok_no") or "",
                l.get("adet") if l.get("adet") is not None else "",
                l.get("detay") or "",
            ]
            for l in loglar
        ],
    )

    ws3 = wb.create_sheet("Özet")
    _sayfa_yaz(ws3, ["Başlık", "Değer"], ozet)
    return wb


def kitap_baytlari(wb):
    tampon = io.BytesIO()
    wb.save(tampon)
    tampon.seek(0)
    return tampon


def dosya_adi(on_ek):
    return f"{on_ek}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
