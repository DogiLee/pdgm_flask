"""
PDGM veri deposu - Excel tabanlı güvenli sürüm.

Bu sürüm önceki uygulamadaki temel sorunları azaltır:
- read/modify/write operasyonları tek RLock altında atomik yapılır.
- dışarı gerçek global dict referansı verilmez; kopya döndürülür.
- kart + log gibi birlikte değişmesi gereken dosyalar önce temp dosyalara hazırlanır,
  ardından best-effort rollback ile birlikte değiştirilir.
- import işlemi önce tamamen doğrulanır, sonra tek kritik bölümde uygulanır.
- internal Excel dosyaları schema ve invariant kontrolüyle okunur.

Sınır: Excel gerçek bir transactional database değildir. Bu tasarım TEK Python process'i
ve birden fazla thread için tasarlanmıştır. Aynı data klasörünü birden fazla process
paylaşmamalıdır. Tam transaction / multi-process güvenliği için SQL'e geçilmelidir.
"""

from __future__ import annotations

import copy
import os
import shutil
import threading
from datetime import date, datetime
from uuid import uuid4

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

KOK = os.path.dirname(os.path.abspath(__file__))
VERI_KLASORU = os.path.join(KOK, "data")
KARTLAR_DOSYA = os.path.join(VERI_KLASORU, "kartlar.xlsx")
LOG_DOSYA = os.path.join(VERI_KLASORU, "islem_logu.xlsx")
YUKLEME_DOSYA = os.path.join(VERI_KLASORU, "yuklemeler.xlsx")
YEDEK_KLASORU = os.path.join(VERI_KLASORU, "yedekler")
PROCESS_KILIT = os.path.join(VERI_KLASORU, "sunucu.lock")

SIRADA = "SIRADA"
DEVAM = "DEVAM EDİYOR"
TAMAM = "TAMAMLANDI"
GECERLI_DURUMLAR = {SIRADA, DEVAM, TAMAM}


class DepoHatasi(Exception):
    """Depo katmanı temel exception sınıfı."""


class KartBulunamadi(DepoHatasi):
    pass


class IsKuralHatasi(DepoHatasi):
    pass


class VeriDogrulamaHatasi(DepoHatasi):
    pass


KART_ALANLARI = [
    ("ID", "id"),
    ("Sıra", "sira"),
    ("Talep NO", "talep_no"),
    ("Kart Stok No", "stok_no"),
    ("Talep Sahibi", "talep_sahibi"),
    ("Toplam Adet", "toplam_adet"),
    ("Adet Metni", "adet_metin"),
    ("Plan Haftası", "plan_hafta"),
    ("Plan Başlangıç", "plan_baslama"),
    ("Plan Teslim", "plan_teslim"),
    ("Gerçekleşen Teslim", "gerceklesen_teslim"),
    ("Excel Durumu", "excel_durum"),
    ("PCB", "pcb"),
    ("Durum", "durum"),
    ("Başlangıç Adedi", "baslangic_adet"),
    ("Tamamlanan Adet", "tamamlanan_adet"),
    ("Başlama Zamanı", "baslama_zamani"),
    ("Bitiş Zamanı", "bitis_zamani"),
    ("Operatör", "operator"),
    ("Not", "aciklama"),
    ("Son Güncelleme", "guncelleme"),
    ("Listede", "aktif"),                     # legacy/genel kayıt yaşam döngüsü
    ("Kaynakta Aktif", "source_active"),      # son Excel batch'inde görüldü mü?
    ("Admin Gizli", "admin_gizli"),           # sonraki import bunu geri açmaz
    ("Kaynak", "kaynak"),
    ("Anahtar", "anahtar"),
]
LOG_ALANLARI = [
    ("Zaman", "zaman"),
    ("Kullanıcı", "kullanici"),
    ("Rol", "rol"),
    ("İşlem", "islem"),
    ("Talep NO", "talep_no"),
    ("Kart Stok No", "stok_no"),
    ("Adet", "adet"),
    ("Detay", "detay"),
]
YUKLEME_ALANLARI = [
    ("Zaman", "zaman"),
    ("Kullanıcı", "kullanici"),
    ("Dosya", "dosya"),
    ("Okunan Satır", "satir"),
    ("Yeni Kart", "yeni"),
    ("Güncellenen", "guncellenen"),
    ("Pasife Alınan", "pasife_alinan"),
    ("Uyarı", "uyari"),
]

SAYISAL = {
    "id",
    "sira",
    "toplam_adet",
    "baslangic_adet",
    "tamamlanan_adet",
    "aktif",
    "source_active",
    "admin_gizli",
    "adet",
    "satir",
    "yeni",
    "guncellenen",
    "pasife_alinan",
    "uyari",
}

ZORUNLU_KART_ALANLARI = {
    "id",
    "talep_no",
    "stok_no",
    "toplam_adet",
    "tamamlanan_adet",
    "durum",
    "anahtar",
}

_kilit = threading.RLock()
_kartlar: list[dict] = []
_loglar: list[dict] = []
_yuklemeler: list[dict] = []

BASLIK_DOLGU = PatternFill("solid", fgColor="0F2027")
BASLIK_YAZI = Font(name="Arial", bold=True, color="FFFFFF", size=11)
GOVDE_YAZI = Font(name="Arial", size=10)

LOG_SINIRI = 20_000
LOG_SAKLA = 5_000


# ------------------------------------------------------------------ yardımcılar

def simdi() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def tarih_coz(deger):
    """Excel'den gelen tarihi YYYY-MM-DD metnine çevirir."""
    if deger is None or deger == "":
        return None
    if isinstance(deger, datetime):
        return deger.strftime("%Y-%m-%d")
    if isinstance(deger, date):
        return deger.strftime("%Y-%m-%d")
    metin = str(deger).strip()
    for kalip in (
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d",
        "%d.%m.%Y",
        "%d/%m/%Y",
        "%Y/%m/%d",
        "%d-%m-%Y",
    ):
        try:
            return datetime.strptime(metin, kalip).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None


def gun_farki(a, b):
    """a - b gün farkı. Metin tarihleri kabul eder."""
    if not a or not b:
        return None
    try:
        t1 = datetime.strptime(str(a)[:10], "%Y-%m-%d").date()
        t2 = datetime.strptime(str(b)[:10], "%Y-%m-%d").date()
    except ValueError:
        return None
    return (t1 - t2).days


def _sayi(deger, varsayilan=0):
    try:
        return int(float(deger))
    except (TypeError, ValueError):
        return varsayilan


def _kart_ref(kart_id: int):
    for kart in _kartlar:
        if kart.get("id") == kart_id:
            return kart
    return None


def _gorunur_mu(kart: dict) -> bool:
    if kart.get("aktif", 1) != 1:
        return False
    if kart.get("admin_gizli", 0) == 1:
        return False
    # Açık işler son Excel batch'inde olmasa bile operatör/pano'da kalır.
    # Aksi halde yarım kalan üretim ekrandan kaybolur.
    if kart.get("durum") in (SIRADA, DEVAM):
        return True
    # Tamamlanan tarihsel kartlar yeni Excel'de artık bulunmasa bile tarihçe için görünür kalır.
    return True


# ------------------------------------------------------------- dosya okuma/yazma

def _oku(dosya, alanlar, zorunlu_alanlar=frozenset()):
    if not os.path.exists(dosya):
        return []

    try:
        wb = openpyxl.load_workbook(dosya, data_only=True, read_only=True)
    except Exception as exc:  # noqa: BLE001
        raise VeriDogrulamaHatasi(
            f"'{os.path.basename(dosya)}' açılamadı: {exc}"
        ) from exc

    try:
        ws = wb[wb.sheetnames[0]]
        satirlar = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()

    if not satirlar:
        return []

    basliklar = [str(h or "").strip() for h in satirlar[0]]
    yerlesim = {}
    for baslik, alan in alanlar:
        if baslik in basliklar:
            yerlesim[alan] = basliklar.index(baslik)

    eksik = sorted(alan for alan in zorunlu_alanlar if alan not in yerlesim)
    if eksik:
        ters = {alan: baslik for baslik, alan in alanlar}
        adlar = ", ".join(ters.get(a, a) for a in eksik)
        raise VeriDogrulamaHatasi(
            f"'{os.path.basename(dosya)}' zorunlu sütunları eksik: {adlar}"
        )

    kayitlar = []
    for satir in satirlar[1:]:
        if not any(h not in (None, "") for h in satir):
            continue
        kayit = {}
        for _, alan in alanlar:
            i = yerlesim.get(alan)
            deger = satir[i] if i is not None and i < len(satir) else None
            if alan in SAYISAL:
                kayit[alan] = _sayi(deger, 0) if deger not in (None, "") else None
            elif isinstance(deger, datetime):
                kayit[alan] = deger.strftime("%Y-%m-%d %H:%M:%S")
            else:
                kayit[alan] = str(deger).strip() if deger not in (None, "") else None
        kayitlar.append(kayit)
    return kayitlar


def _workbook_uret(alanlar, kayitlar, sayfa_adi):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sayfa_adi

    ws.append([baslik for baslik, _ in alanlar])
    for hucre in ws[1]:
        hucre.fill = BASLIK_DOLGU
        hucre.font = BASLIK_YAZI
        hucre.alignment = Alignment(horizontal="center", vertical="center")

    for kayit in kayitlar:
        ws.append([kayit.get(alan) for _, alan in alanlar])

    for satir in ws.iter_rows(min_row=2):
        for hucre in satir:
            hucre.font = GOVDE_YAZI

    for sutun, (baslik, alan) in enumerate(alanlar, start=1):
        en = max([len(baslik)] + [len(str(k.get(alan) or "")) for k in kayitlar[:300]] or [10])
        ws.column_dimensions[get_column_letter(sutun)].width = min(40, max(11, en + 3))
    ws.freeze_panes = "A2"
    return wb


def _temp_yaz(hedef, alanlar, kayitlar, sayfa_adi):
    os.makedirs(os.path.dirname(hedef), exist_ok=True)
    temp = f"{hedef}.{uuid4().hex}.yeni"
    wb = _workbook_uret(alanlar, kayitlar, sayfa_adi)
    try:
        wb.save(temp)
    finally:
        wb.close()
    return temp


def _coklu_yaz(dosyalar):
    """
    Birden fazla workbook'u önce temp dosyalara hazırlar, sonra hedeflerle değiştirir.

    `dosyalar`: [(hedef, alanlar, kayitlar, sayfa_adi), ...]

    Gerçek DB transaction'ı değildir; ancak normal Python exception / file-lock hatalarında
    daha önce değiştirilmiş hedefleri backup'tan geri yüklemeyi dener.
    """
    os.makedirs(VERI_KLASORU, exist_ok=True)
    temps = []
    backups = {}
    degisen = []

    try:
        for hedef, alanlar, kayitlar, sayfa_adi in dosyalar:
            temps.append((hedef, _temp_yaz(hedef, alanlar, kayitlar, sayfa_adi)))

        for hedef, _ in temps:
            if os.path.exists(hedef):
                backup = f"{hedef}.{uuid4().hex}.txn.bak"
                shutil.copy2(hedef, backup)
                backups[hedef] = backup
            else:
                backups[hedef] = None

        for hedef, temp in temps:
            os.replace(temp, hedef)
            degisen.append(hedef)

    except Exception:
        # Best-effort rollback.
        for hedef in reversed(degisen):
            backup = backups.get(hedef)
            try:
                if backup and os.path.exists(backup):
                    os.replace(backup, hedef)
                elif os.path.exists(hedef):
                    os.remove(hedef)
            except OSError:
                pass
        raise
    finally:
        for _, temp in temps:
            if os.path.exists(temp):
                try:
                    os.remove(temp)
                except OSError:
                    pass
        for backup in backups.values():
            if backup and os.path.exists(backup):
                try:
                    os.remove(backup)
                except OSError:
                    pass


def _yaz(dosya, alanlar, kayitlar, sayfa_adi):
    _coklu_yaz([(dosya, alanlar, kayitlar, sayfa_adi)])


def _gunluk_yedek(dosya):
    if not os.path.exists(dosya):
        return
    os.makedirs(YEDEK_KLASORU, exist_ok=True)
    hedef = os.path.join(YEDEK_KLASORU, f"{date.today():%Y%m%d}_{os.path.basename(dosya)}")
    if not os.path.exists(hedef):
        shutil.copy2(dosya, hedef)


def anlik_yedek(etiket: str = "once") -> str:
    """
    kartlar / log / yüklemeler dosyalarının timestamp'li kopyasını alır.
    Import ve kritik bakım öncesi çağrılır; geri dönüş için klasör yolu döner.
    """
    os.makedirs(YEDEK_KLASORU, exist_ok=True)
    damga = datetime.now().strftime("%Y%m%d_%H%M%S")
    guvenli = "".join(ch if ch.isalnum() or ch in "-_" else "_" for ch in (etiket or "yedek"))
    klasor = os.path.join(YEDEK_KLASORU, f"{damga}_{guvenli}")
    os.makedirs(klasor, exist_ok=True)
    for dosya in (KARTLAR_DOSYA, LOG_DOSYA, YUKLEME_DOSYA):
        if os.path.exists(dosya):
            shutil.copy2(dosya, os.path.join(klasor, os.path.basename(dosya)))
    return klasor


def _pid_yasiyor_mu(pid: int) -> bool:
    if pid <= 0:
        return False
    try:
        os.kill(pid, 0)
    except ProcessLookupError:
        return False
    except PermissionError:
        return True
    except OSError:
        return False
    return True


def process_kilidi_al():
    """
    Aynı data klasörünü ikinci bir Flask/Waitress process'inin açmasını engeller.
    Stale lock (çökmüş process) otomatik temizlenir.
    """
    os.makedirs(VERI_KLASORU, exist_ok=True)
    if os.path.exists(PROCESS_KILIT):
        try:
            with open(PROCESS_KILIT, encoding="utf-8") as f:
                eski = int((f.read() or "0").strip() or "0")
        except (OSError, ValueError):
            eski = 0
        if eski and _pid_yasiyor_mu(eski) and eski != os.getpid():
            raise RuntimeError(
                f"data/ klasörü zaten başka bir PDGM process tarafından kullanılıyor "
                f"(PID {eski}). İkinci sunucu açmayın; Excel bozulabilir. "
                f"Eski process kapalıysa '{PROCESS_KILIT}' dosyasını silip tekrar deneyin."
            )
        if eski != os.getpid():
            try:
                os.remove(PROCESS_KILIT)
            except OSError:
                pass

    if os.path.exists(PROCESS_KILIT):
        try:
            with open(PROCESS_KILIT, encoding="utf-8") as f:
                if (f.read() or "").strip() == str(os.getpid()):
                    return
        except OSError:
            pass

    with open(PROCESS_KILIT, "w", encoding="utf-8") as f:
        f.write(str(os.getpid()))

    if getattr(process_kilidi_al, "_atexit_bagli", False):
        return

    def _birak():
        try:
            if not os.path.exists(PROCESS_KILIT):
                return
            with open(PROCESS_KILIT, encoding="utf-8") as f:
                sahip = (f.read() or "").strip()
            if sahip == str(os.getpid()):
                os.remove(PROCESS_KILIT)
        except OSError:
            pass

    import atexit

    atexit.register(_birak)
    process_kilidi_al._atexit_bagli = True



# --------------------------------------------------------------- doğrulamalar

def _kart_normalize(kart: dict) -> dict:
    kart = dict(kart)
    kart["toplam_adet"] = kart.get("toplam_adet") or 1
    kart["baslangic_adet"] = kart.get("baslangic_adet") or 0
    kart["tamamlanan_adet"] = kart.get("tamamlanan_adet") or 0
    kart["aktif"] = 1 if kart.get("aktif") in (None, 1) else 0
    kart["source_active"] = 1 if kart.get("source_active") in (None, 1) else 0
    kart["admin_gizli"] = 1 if kart.get("admin_gizli") == 1 else 0
    kart["kaynak"] = kart.get("kaynak") or "EXCEL"
    kart["durum"] = kart.get("durum") or SIRADA
    return kart


def _kart_dogrula(kart: dict):
    kart_id = _sayi(kart.get("id"), -1)
    toplam = _sayi(kart.get("toplam_adet"), 0)
    tamam = _sayi(kart.get("tamamlanan_adet"), 0)
    durum = kart.get("durum")

    if kart_id < 1:
        raise VeriDogrulamaHatasi("Kart ID pozitif tam sayı olmalı.")
    if toplam < 1:
        raise VeriDogrulamaHatasi(f"Kart {kart_id}: Toplam Adet en az 1 olmalı.")
    if tamam < 0 or tamam > toplam:
        raise VeriDogrulamaHatasi(
            f"Kart {kart_id}: Tamamlanan Adet ({tamam}) 0 ile Toplam Adet ({toplam}) arasında olmalı."
        )
    if durum not in GECERLI_DURUMLAR:
        raise VeriDogrulamaHatasi(f"Kart {kart_id}: Geçersiz durum '{durum}'.")
    if durum == TAMAM and tamam != toplam:
        raise VeriDogrulamaHatasi(
            f"Kart {kart_id}: TAMAMLANDI durumunda tamamlanan adet toplam adede eşit olmalı."
        )
    if durum == DEVAM and tamam >= toplam:
        raise VeriDogrulamaHatasi(
            f"Kart {kart_id}: DEVAM durumunda tamamlanan adet toplam adetten küçük olmalı."
        )
    if durum == SIRADA and tamam != 0:
        raise VeriDogrulamaHatasi(
            f"Kart {kart_id}: SIRADA durumunda tamamlanan adet 0 olmalı."
        )
    if not kart.get("anahtar"):
        raise VeriDogrulamaHatasi(f"Kart {kart_id}: Anahtar boş olamaz.")


def _kart_listesi_dogrula(kartlar):
    idler = set()
    anahtarlar = set()
    for kart in kartlar:
        _kart_dogrula(kart)
        if kart["id"] in idler:
            raise VeriDogrulamaHatasi(f"Tekrarlanan kart ID: {kart['id']}")
        if kart["anahtar"] in anahtarlar:
            raise VeriDogrulamaHatasi(f"Tekrarlanan kart anahtarı: {kart['anahtar']}")
        idler.add(kart["id"])
        anahtarlar.add(kart["anahtar"])


# ------------------------------------------------------------------- başlangıç

def kur():
    global _kartlar, _loglar, _yuklemeler
    with _kilit:
        os.makedirs(VERI_KLASORU, exist_ok=True)

        kartlar = _oku(
            KARTLAR_DOSYA,
            KART_ALANLARI,
            ZORUNLU_KART_ALANLARI if os.path.exists(KARTLAR_DOSYA) else frozenset(),
        )
        kartlar = [_kart_normalize(k) for k in kartlar]
        if kartlar:
            _kart_listesi_dogrula(kartlar)

        loglar = _oku(LOG_DOSYA, LOG_ALANLARI)
        yuklemeler = _oku(YUKLEME_DOSYA, YUKLEME_ALANLARI)

        _kartlar = kartlar
        _loglar = loglar
        _yuklemeler = yuklemeler

        eksikler = []
        if not os.path.exists(KARTLAR_DOSYA):
            eksikler.append((KARTLAR_DOSYA, KART_ALANLARI, _kartlar, "Kartlar"))
        if not os.path.exists(LOG_DOSYA):
            eksikler.append((LOG_DOSYA, LOG_ALANLARI, _loglar, "İşlem Logu"))
        if not os.path.exists(YUKLEME_DOSYA):
            eksikler.append((YUKLEME_DOSYA, YUKLEME_ALANLARI, _yuklemeler, "Yüklemeler"))
        if eksikler:
            _coklu_yaz(eksikler)


def kartlari_diskten_yeniden_yukle():
    """Sadece kartlar.xlsx'i doğrulayarak yeniden yükler; log/yükleme geçmişine dokunmaz."""
    global _kartlar
    with _kilit:
        yeni = _oku(KARTLAR_DOSYA, KART_ALANLARI, ZORUNLU_KART_ALANLARI)
        yeni = [_kart_normalize(k) for k in yeni]
        _kart_listesi_dogrula(yeni)
        _kartlar = yeni
        return len(_kartlar)


def yeniden_yukle():
    """Geriye uyumluluk: artık yalnız kart dosyasını yeniden yükler."""
    return kartlari_diskten_yeniden_yukle()


def _kartlari_kaydet():
    _gunluk_yedek(KARTLAR_DOSYA)
    _yaz(KARTLAR_DOSYA, KART_ALANLARI, _kartlar, "Kartlar")


# ------------------------------------------------------------------------- log

def _log_kaydi(kullanici, rol, islem, talep_no="", stok_no="", adet=None, detay=""):
    return {
        "zaman": simdi(),
        "kullanici": kullanici,
        "rol": rol,
        "islem": islem,
        "talep_no": talep_no,
        "stok_no": stok_no,
        "adet": adet,
        "detay": detay,
    }


def _log_arsivle_gerekirse():
    global _loglar
    if len(_loglar) <= LOG_SINIRI:
        return

    os.makedirs(YEDEK_KLASORU, exist_ok=True)
    arsivlenecek = _loglar[:-LOG_SAKLA]
    kalan = _loglar[-LOG_SAKLA:]
    arsiv = os.path.join(YEDEK_KLASORU, f"{datetime.now():%Y%m%d_%H%M%S}_islem_logu_arsiv.xlsx")
    _yaz(arsiv, LOG_ALANLARI, arsivlenecek, "İşlem Logu")
    _loglar = kalan


def log_ekle(kullanici, rol, islem, talep_no="", stok_no="", adet=None, detay=""):
    global _loglar
    with _kilit:
        eski = copy.deepcopy(_loglar)
        try:
            _loglar.append(_log_kaydi(kullanici, rol, islem, talep_no, stok_no, adet, detay))
            _log_arsivle_gerekirse()
            _yaz(LOG_DOSYA, LOG_ALANLARI, _loglar, "İşlem Logu")
        except Exception:
            _loglar = eski
            raise


def loglari_getir(adet=None):
    with _kilit:
        ters = copy.deepcopy(list(reversed(_loglar)))
        return ters[:adet] if adet else ters


def yukleme_ekle(kullanici, dosya, satir, yeni, guncellenen, pasife_alinan=0, uyari=0):
    global _yuklemeler
    with _kilit:
        eski = copy.deepcopy(_yuklemeler)
        try:
            _yuklemeler.append(
                {
                    "zaman": simdi(),
                    "kullanici": kullanici,
                    "dosya": dosya,
                    "satir": satir,
                    "yeni": yeni,
                    "guncellenen": guncellenen,
                    "pasife_alinan": pasife_alinan,
                    "uyari": uyari,
                }
            )
            _yaz(YUKLEME_DOSYA, YUKLEME_ALANLARI, _yuklemeler, "Yüklemeler")
        except Exception:
            _yuklemeler = eski
            raise


def yuklemeleri_getir(adet=None):
    with _kilit:
        ters = copy.deepcopy(list(reversed(_yuklemeler)))
        return ters[:adet] if adet else ters


# ------------------------------------------------------------- durum ve sapma

def durum_bilgisi(k):
    bugun = date.today().strftime("%Y-%m-%d")
    durum = k.get("durum") or SIRADA
    plan_teslim = k.get("plan_teslim")
    plan_baslama = k.get("plan_baslama")

    bilgi = {
        "rozet": "SIRADA",
        "renk": "notr",
        "sapma": None,
        "kalan": None,
        "zaman_yuzde": 0,
        "plan_gun": gun_farki(plan_teslim, plan_baslama),
    }

    if durum == TAMAM:
        bitis = str(k.get("bitis_zamani") or k.get("gerceklesen_teslim") or "")[:10]
        sapma = gun_farki(bitis, plan_teslim)
        bilgi["sapma"] = sapma
        bilgi["zaman_yuzde"] = 100
        if sapma is None:
            bilgi["rozet"], bilgi["renk"] = "TAMAMLANDI", "iyi"
        elif sapma > 0:
            bilgi["rozet"], bilgi["renk"] = f"GECİKMELİ BİTTİ (+{sapma} gün)", "kotu"
        else:
            bilgi["rozet"], bilgi["renk"] = "ZAMANINDA BİTTİ", "iyi"
        return bilgi

    if durum == DEVAM:
        kalan = gun_farki(plan_teslim, bugun)
        bilgi["kalan"] = kalan
        gecen = gun_farki(bugun, str(k.get("baslama_zamani") or plan_baslama or bugun)[:10]) or 0
        plan_gun = bilgi["plan_gun"]
        if plan_gun and plan_gun > 0:
            bilgi["zaman_yuzde"] = max(0, min(140, round(gecen / plan_gun * 100)))
        else:
            bilgi["zaman_yuzde"] = 100 if kalan is not None and kalan < 0 else 50
        if kalan is None:
            bilgi["rozet"], bilgi["renk"] = "DEVAM EDİYOR", "uyari"
        elif kalan < 0:
            bilgi["sapma"] = -kalan
            bilgi["rozet"], bilgi["renk"] = f"SÜRE AŞILDI ({-kalan} gün)", "kotu"
        elif kalan <= 1:
            bilgi["rozet"] = "SON GÜN" if kalan == 0 else "SON 1 GÜN"
            bilgi["renk"] = "uyari"
        else:
            bilgi["rozet"], bilgi["renk"] = f"PLANINDA ({kalan} gün var)", "iyi"
        return bilgi

    gecikme = gun_farki(bugun, plan_baslama)
    if plan_baslama and gecikme is not None and gecikme > 0:
        bilgi["rozet"], bilgi["renk"], bilgi["sapma"] = (
            f"BAŞLAMADI (+{gecikme} gün)",
            "kotu",
            gecikme,
        )
    elif plan_baslama and gecikme == 0:
        bilgi["rozet"], bilgi["renk"] = "BUGÜN BAŞLAMALI", "uyari"
    return bilgi


def kart_gorunumu(k):
    d = copy.deepcopy(k)
    d.update(durum_bilgisi(k))
    d["toplam_adet"] = d.get("toplam_adet") or 1
    d["tamamlanan_adet"] = d.get("tamamlanan_adet") or 0
    d["baslangic_adet"] = d.get("baslangic_adet") or 0
    d["kalan_adet"] = max(0, d["toplam_adet"] - d["tamamlanan_adet"])
    d["adet_yuzde"] = min(100, round(d["tamamlanan_adet"] / d["toplam_adet"] * 100))
    d["gorunur"] = _gorunur_mu(d)
    d["kaynakta_yok"] = int(d.get("source_active", 1) or 0) != 1
    d["is_durumu"] = d.get("durum") or SIRADA
    d["kaynak_durumu"] = (d.get("excel_durum") or "").strip()
    return d


SIRALAMA = {DEVAM: 0, SIRADA: 1, TAMAM: 2}


def kartlari_getir(sadece_gorunen=True):
    with _kilit:
        secim = [k for k in _kartlar if not sadece_gorunen or _gorunur_mu(k)]
        kartlar = [kart_gorunumu(k) for k in secim]
    kartlar.sort(
        key=lambda k: (
            SIRALAMA.get(k["durum"], 3),
            k.get("plan_baslama") or "9999-12-31",
            k.get("sira") or 9999,
        )
    )
    return kartlar


def kart_getir(kart_id):
    kart_id = _sayi(kart_id, -1)
    with _kilit:
        kart = _kart_ref(kart_id)
        return kart_gorunumu(kart) if kart else None


def kart_bul(anahtar):
    """Geriye uyumluluk: gerçek global dict yerine kopya döndürür."""
    with _kilit:
        for k in _kartlar:
            if k.get("anahtar") == anahtar:
                return copy.deepcopy(k)
    return None


def yeni_kimlik():
    with _kilit:
        return max([_sayi(k.get("id"), 0) for k in _kartlar] or [0]) + 1


# ----------------------------------------------------------- atomik kart işlemleri

def _kart_log_commit(eski_kartlar, eski_loglar):
    """Kart + log dosyasını birlikte yaz; hata halinde RAM'i çağıran geri yükler."""
    _gunluk_yedek(KARTLAR_DOSYA)
    _coklu_yaz(
        [
            (KARTLAR_DOSYA, KART_ALANLARI, _kartlar, "Kartlar"),
            (LOG_DOSYA, LOG_ALANLARI, _loglar, "İşlem Logu"),
        ]
    )


def kart_baslat(kart_id, adet, kullanici, rol, aciklama=""):
    global _kartlar, _loglar
    kart_id = _sayi(kart_id, -1)

    with _kilit:
        kart = _kart_ref(kart_id)
        if not kart:
            raise KartBulunamadi("Kart bulunamadı.")
        if not _gorunur_mu(kart):
            raise IsKuralHatasi("Bu kart aktif listede değil.")
        if kart["durum"] != SIRADA:
            raise IsKuralHatasi("Yalnız SIRADA durumundaki kart başlatılabilir.")

        kalan = kart["toplam_adet"] - kart["tamamlanan_adet"]
        if adet in (None, ""):
            adet = kalan
        try:
            adet = int(adet)
        except (TypeError, ValueError) as exc:
            raise ValueError("Adet sayı olmalı.") from exc
        if adet < 1 or adet > kalan:
            raise IsKuralHatasi(f"Adet 1 ile {kalan} arasında olmalı.")

        eski_kartlar = copy.deepcopy(_kartlar)
        eski_loglar = copy.deepcopy(_loglar)
        try:
            kart.update(
                durum=DEVAM,
                baslangic_adet=adet,
                baslama_zamani=simdi(),
                operator=kullanici,
                aciklama=aciklama or kart.get("aciklama"),
                guncelleme=simdi(),
            )
            _loglar.append(
                _log_kaydi(
                    kullanici,
                    rol,
                    "İŞE BAŞLANDI",
                    kart.get("talep_no") or "",
                    kart.get("stok_no") or "",
                    adet,
                    aciklama or f"{adet} adet dizgiye alındı",
                )
            )
            _kart_dogrula(kart)
            _kart_log_commit(eski_kartlar, eski_loglar)
        except Exception:
            _kartlar = eski_kartlar
            _loglar = eski_loglar
            raise

        return kart_gorunumu(kart)


def kart_bitir(kart_id, adet, kullanici, rol, aciklama=""):
    global _kartlar, _loglar
    kart_id = _sayi(kart_id, -1)

    with _kilit:
        kart = _kart_ref(kart_id)
        if not kart:
            raise KartBulunamadi("Kart bulunamadı.")
        if kart["durum"] != DEVAM:
            raise IsKuralHatasi("Önce kartı başlatın.")

        kalan = kart["toplam_adet"] - kart["tamamlanan_adet"]
        try:
            adet = int(adet)
        except (TypeError, ValueError) as exc:
            raise ValueError("Adet sayı olmalı.") from exc
        if adet < 1 or adet > kalan:
            raise IsKuralHatasi(f"Adet 1 ile {kalan} arasında olmalı.")

        eski_kartlar = copy.deepcopy(_kartlar)
        eski_loglar = copy.deepcopy(_loglar)
        try:
            yeni_toplam = kart["tamamlanan_adet"] + adet
            bitti = yeni_toplam == kart["toplam_adet"]
            kart.update(
                durum=TAMAM if bitti else DEVAM,
                tamamlanan_adet=yeni_toplam,
                bitis_zamani=simdi() if bitti else None,
                operator=kullanici,
                aciklama=aciklama or kart.get("aciklama"),
                guncelleme=simdi(),
            )
            if bitti:
                kart["baslangic_adet"] = 0

            _loglar.append(
                _log_kaydi(
                    kullanici,
                    rol,
                    "TAMAMLANDI" if bitti else "KISMİ TESLİM",
                    kart.get("talep_no") or "",
                    kart.get("stok_no") or "",
                    adet,
                    aciklama or f"{yeni_toplam}/{kart['toplam_adet']} adet bitti",
                )
            )
            _kart_dogrula(kart)
            _kart_log_commit(eski_kartlar, eski_loglar)
        except Exception:
            _kartlar = eski_kartlar
            _loglar = eski_loglar
            raise

        mesaj = (
            "Kart tamamlandı."
            if bitti
            else f"{yeni_toplam}/{kart['toplam_adet']} adet bitti, kart devam ediyor."
        )
        return kart_gorunumu(kart), bitti, mesaj


def kart_not_guncelle(kart_id, aciklama, kullanici, rol):
    global _kartlar, _loglar
    kart_id = _sayi(kart_id, -1)
    with _kilit:
        kart = _kart_ref(kart_id)
        if not kart:
            raise KartBulunamadi("Kart bulunamadı.")

        eski_kartlar = copy.deepcopy(_kartlar)
        eski_loglar = copy.deepcopy(_loglar)
        try:
            kart["aciklama"] = aciklama
            kart["guncelleme"] = simdi()
            _loglar.append(
                _log_kaydi(
                    kullanici,
                    rol,
                    "NOT GÜNCELLENDİ",
                    kart.get("talep_no") or "",
                    kart.get("stok_no") or "",
                    None,
                    aciklama,
                )
            )
            _kart_log_commit(eski_kartlar, eski_loglar)
        except Exception:
            _kartlar = eski_kartlar
            _loglar = eski_loglar
            raise
        return kart_gorunumu(kart)


def admin_kart_duzenle(kart_id, durum, tamamlanan_adet, toplam_adet, aciklama, kullanici):
    global _kartlar, _loglar
    kart_id = _sayi(kart_id, -1)

    with _kilit:
        kart = _kart_ref(kart_id)
        if not kart:
            raise KartBulunamadi("Kart bulunamadı.")

        yeni_durum = durum or kart["durum"]
        if yeni_durum not in GECERLI_DURUMLAR:
            raise IsKuralHatasi("Geçersiz durum.")

        try:
            toplam = int(kart["toplam_adet"] if toplam_adet in (None, "") else toplam_adet)
            tamamlanan = int(
                kart["tamamlanan_adet"] if tamamlanan_adet in (None, "") else tamamlanan_adet
            )
        except (TypeError, ValueError) as exc:
            raise ValueError("Adetler sayı olmalı.") from exc

        if toplam < 1 or tamamlanan < 0 or tamamlanan > toplam:
            raise IsKuralHatasi("Adet değerleri tutarsız.")

        if yeni_durum == TAMAM:
            tamamlanan = toplam
            bitis = kart.get("bitis_zamani") or simdi()
            baslama = kart.get("baslama_zamani") or simdi()
        elif yeni_durum == DEVAM:
            if tamamlanan >= toplam:
                raise IsKuralHatasi("DEVAM durumunda tamamlanan adet toplam adetten küçük olmalı.")
            bitis = None
            baslama = kart.get("baslama_zamani") or simdi()
        else:
            if tamamlanan != 0:
                raise IsKuralHatasi("SIRADA durumunda tamamlanan adet 0 olmalı.")
            bitis = None
            baslama = None

        eski_kartlar = copy.deepcopy(_kartlar)
        eski_loglar = copy.deepcopy(_loglar)
        onceki = kart["durum"]
        try:
            kart.update(
                durum=yeni_durum,
                toplam_adet=toplam,
                tamamlanan_adet=tamamlanan,
                baslama_zamani=baslama,
                bitis_zamani=bitis,
                aciklama=aciklama if aciklama is not None else kart.get("aciklama"),
                guncelleme=simdi(),
            )
            if yeni_durum == SIRADA:
                kart["baslangic_adet"] = 0
            _kart_dogrula(kart)
            _loglar.append(
                _log_kaydi(
                    kullanici,
                    "admin",
                    "ADMİN DÜZENLEDİ",
                    kart.get("talep_no") or "",
                    kart.get("stok_no") or "",
                    tamamlanan,
                    f"{onceki} → {yeni_durum} · {tamamlanan}/{toplam} adet",
                )
            )
            _kart_log_commit(eski_kartlar, eski_loglar)
        except Exception:
            _kartlar = eski_kartlar
            _loglar = eski_loglar
            raise

        return kart_gorunumu(kart)


def admin_kart_gizle(kart_id, kullanici):
    global _kartlar, _loglar
    kart_id = _sayi(kart_id, -1)
    with _kilit:
        kart = _kart_ref(kart_id)
        if not kart:
            raise KartBulunamadi("Kart bulunamadı.")

        eski_kartlar = copy.deepcopy(_kartlar)
        eski_loglar = copy.deepcopy(_loglar)
        try:
            kart["admin_gizli"] = 1
            kart["guncelleme"] = simdi()
            _loglar.append(
                _log_kaydi(
                    kullanici,
                    "admin",
                    "KART LİSTEDEN GİZLENDİ",
                    kart.get("talep_no") or "",
                    kart.get("stok_no") or "",
                    None,
                    "Admin tarafından gizlendi; sonraki Excel importu otomatik geri açmaz.",
                )
            )
            _kart_log_commit(eski_kartlar, eski_loglar)
        except Exception:
            _kartlar = eski_kartlar
            _loglar = eski_loglar
            raise


# -------------------------------------------------------------- Excel import commit

def excel_import_uygula(dosya_adi, kullanici, satirlar, uyari_sayisi=0):
    """
    excel_araclari tarafından tamamen parse/validate edilmiş satırları tek kritik bölümde uygular.

    `satirlar` elemanları:
      {
        "anahtar": str,
        "talep_no": str,
        "stok_no": str,
        "plan": {...},
        "gerceklesen_teslim": str|None,
        "ilk_durum": SIRADA|DEVAM|TAMAM,
      }
    """
    global _kartlar, _loglar, _yuklemeler

    with _kilit:
        eski_kartlar = copy.deepcopy(_kartlar)
        eski_loglar = copy.deepcopy(_loglar)
        eski_yuklemeler = copy.deepcopy(_yuklemeler)

        try:
            yedek_klasoru = anlik_yedek("import_oncesi")
            mevcut_harita = {k.get("anahtar"): k for k in _kartlar}
            gorulen = set()
            yeni = 0
            guncellenen = 0
            workflow_korundu = 0
            pasife_listesi = []
            sonraki_id = max([_sayi(k.get("id"), 0) for k in _kartlar] or [0]) + 1

            for satir in satirlar:
                anahtar = satir["anahtar"]
                gorulen.add(anahtar)
                plan = satir["plan"]
                mevcut = mevcut_harita.get(anahtar)

                if mevcut:
                    yeni_toplam = int(plan["toplam_adet"])
                    tamamlanan = int(mevcut.get("tamamlanan_adet") or 0)
                    if mevcut.get("durum") == TAMAM and yeni_toplam != tamamlanan:
                        raise IsKuralHatasi(
                            f"{anahtar}: Kart sistemde TAMAMLANDI ({tamamlanan}/{tamamlanan}). "
                            f"Yeni Excel toplam adedi {yeni_toplam}. Tamamlanmış kartın toplam adedi "
                            "otomatik değiştirilemez; admin kontrolü gerekir. Import iptal edildi."
                        )
                    if mevcut.get("durum") == DEVAM and yeni_toplam <= tamamlanan:
                        raise IsKuralHatasi(
                            f"{anahtar}: Kart DEVAM durumunda ve {tamamlanan} adet tamamlanmış. "
                            f"Yeni Excel toplam adedi {yeni_toplam}; tamamlanan adetten büyük olmalı. "
                            "Import iptal edildi."
                        )
                    if yeni_toplam < tamamlanan:
                        raise IsKuralHatasi(
                            f"{anahtar}: Excel toplam adedi ({yeni_toplam}), sistemde tamamlanan "
                            f"adetten ({tamamlanan}) küçük olamaz. Import iptal edildi."
                        )

                    # Excel yalnız plan/source alanlarını günceller; web workflow durumunu ezmez.
                    onceki_durum = mevcut.get("durum")
                    onceki_tamam = tamamlanan
                    mevcut.update(plan)
                    mevcut["gerceklesen_teslim"] = satir.get("gerceklesen_teslim")
                    mevcut["source_active"] = 1
                    mevcut["aktif"] = 1
                    mevcut["kaynak"] = "EXCEL"
                    mevcut["guncelleme"] = simdi()
                    # admin_gizli özellikle korunur.
                    # Workflow alanları plan update ile gelmez; yine de açıkça koru.
                    mevcut["durum"] = onceki_durum
                    mevcut["tamamlanan_adet"] = onceki_tamam
                    _kart_dogrula(mevcut)
                    guncellenen += 1
                    workflow_korundu += 1
                    continue

                toplam = int(plan["toplam_adet"])
                ilk_durum = satir["ilk_durum"]
                gerceklesen = satir.get("gerceklesen_teslim")

                tamamlanan = toplam if ilk_durum == TAMAM else 0
                baslangic_adet = toplam if ilk_durum == DEVAM else 0
                baslama = None
                bitis = None

                if ilk_durum in (DEVAM, TAMAM):
                    if plan.get("plan_baslama"):
                        baslama = f"{plan['plan_baslama']} 08:00:00"
                    else:
                        baslama = simdi() if ilk_durum == DEVAM else None

                # Gerçekleşen teslim yoksa plan tarihini uydurma bitiş tarihi olarak kullanma.
                if ilk_durum == TAMAM and gerceklesen:
                    bitis = f"{gerceklesen} 17:00:00"

                kayit = {
                    "id": sonraki_id,
                    "anahtar": anahtar,
                    "talep_no": satir["talep_no"],
                    "stok_no": satir["stok_no"],
                    "gerceklesen_teslim": gerceklesen,
                    "durum": ilk_durum,
                    "baslangic_adet": baslangic_adet,
                    "tamamlanan_adet": tamamlanan,
                    "baslama_zamani": baslama,
                    "bitis_zamani": bitis,
                    "operator": "Excel" if ilk_durum != SIRADA else None,
                    "aciklama": None,
                    "guncelleme": simdi(),
                    "aktif": 1,
                    "source_active": 1,
                    "admin_gizli": 0,
                    "kaynak": "EXCEL",
                }
                kayit.update(plan)
                _kart_dogrula(kayit)
                _kartlar.append(kayit)
                mevcut_harita[anahtar] = kayit
                sonraki_id += 1
                yeni += 1

            # Son batch'te görülmeyen Excel kaynaklı açık kartları source-inactive yap.
            pasife_alinan = 0
            for kart in _kartlar:
                if kart.get("kaynak") != "EXCEL":
                    continue
                if kart.get("anahtar") in gorulen:
                    continue
                if kart.get("durum") == TAMAM:
                    # Tarihsel tamamlanan kayıtlar korunur; kaynakta görünmediği ayrıca işaretlenir.
                    kart["source_active"] = 0
                    continue
                if kart.get("source_active", 1) == 1:
                    kart["source_active"] = 0
                    kart["guncelleme"] = simdi()
                    pasife_alinan += 1
                    pasife_listesi.append(
                        f"{kart.get('talep_no') or ''}|{kart.get('stok_no') or ''}"
                    )

            _kart_listesi_dogrula(_kartlar)

            _yuklemeler.append(
                {
                    "zaman": simdi(),
                    "kullanici": kullanici,
                    "dosya": dosya_adi,
                    "satir": len(satirlar),
                    "yeni": yeni,
                    "guncellenen": guncellenen,
                    "pasife_alinan": pasife_alinan,
                    "uyari": uyari_sayisi,
                }
            )
            _loglar.append(
                _log_kaydi(
                    kullanici,
                    "admin",
                    "EXCEL YÜKLENDİ",
                    detay=(
                        f"{len(satirlar)} satır · {yeni} yeni · {guncellenen} güncellendi · "
                        f"workflow korundu · {pasife_alinan} kaynakta pasif · "
                        f"{uyari_sayisi} uyarı · yedek={os.path.basename(yedek_klasoru)}"
                    ),
                )
            )

            _gunluk_yedek(KARTLAR_DOSYA)
            _coklu_yaz(
                [
                    (KARTLAR_DOSYA, KART_ALANLARI, _kartlar, "Kartlar"),
                    (LOG_DOSYA, LOG_ALANLARI, _loglar, "İşlem Logu"),
                    (YUKLEME_DOSYA, YUKLEME_ALANLARI, _yuklemeler, "Yüklemeler"),
                ]
            )

            return {
                "satir": len(satirlar),
                "yeni": yeni,
                "guncellenen": guncellenen,
                "pasife_alinan": pasife_alinan,
                "pasife_listesi": pasife_listesi[:20],
                "workflow_korundu": workflow_korundu,
                "uyari": uyari_sayisi,
                "yedek": yedek_klasoru,
            }

        except Exception:
            _kartlar = eski_kartlar
            _loglar = eski_loglar
            _yuklemeler = eski_yuklemeler
            raise


# ------------------------------------------------------------- geriye uyumlu API

def kart_guncelle(kart_id, **alanlar):
    """
    Legacy yardımcı. Yeni route'larda kullanılmamalı.
    Tek kart güncellemesini lock altında yapar, fakat log transaction'ı içermez.
    """
    global _kartlar
    kart_id = _sayi(kart_id, -1)
    with _kilit:
        kart = _kart_ref(kart_id)
        if not kart:
            return None
        eski = copy.deepcopy(_kartlar)
        try:
            kart.update(alanlar)
            kart["guncelleme"] = simdi()
            _kart_dogrula(kart)
            _kartlari_kaydet()
            return kart_gorunumu(kart)
        except Exception:
            _kartlar = eski
            raise


def toplu_kaydet(yeni_kartlar=(), degisen=True):
    """Legacy yardımcı; yeni Excel importunda `excel_import_uygula` kullanılmalı."""
    global _kartlar
    with _kilit:
        eski = copy.deepcopy(_kartlar)
        try:
            for kart in yeni_kartlar:
                _kartlar.append(_kart_normalize(kart))
            _kart_listesi_dogrula(_kartlar)
            if degisen:
                _kartlari_kaydet()
        except Exception:
            _kartlar = eski
            raise
